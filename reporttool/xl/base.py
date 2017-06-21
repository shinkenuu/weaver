import abc
from openpyxl import Workbook
from openpyxl import utils
from openpyxl import styles
from openpyxl import formatting


REPORT_ROOT_PATH = '/mnt/jatobrfiles/Weaver/reports/'


def xl(current_index: int):
    return current_index + 1


class Report(metaclass=abc.ABCMeta):
    def __init__(self):
        self.matrix = [[]]
        self.wb = Workbook()
        self.ws = self.wb.active

    class Header(metaclass=abc.ABCMeta):
        def __init__(self, header_name: str, number_format: str, offset: int):
            self.header_name = header_name
            self.number_format = number_format
            self.offset = offset

    @abc.abstractmethod
    def generate_report(self):
        pass

    def write_to_disc(self):
        self.wb.save('{}{}.xlsx'.format(REPORT_ROOT_PATH, str(self).replace('.', '/')))

    def write_matrix_to_xl(self):
        for row in range(len(self.matrix)):
            for col in range(len(self.matrix[row])):
                self.ws.cell(row=row+1, column=col+1).value = self.matrix[row][col]


class EvolutionReport(Report, abc.ABC):
    def __init__(self):
        self.sample_dates = []
        self.sample_headers = []
        self.vehicle_desc_mark_up_col = -1
        super().__init__()

    def __str__(self):
        return 'evolution'

    class Header(Report.Header):
        def __init__(self, header_name: str, number_format: str, offset: int):
            super().__init__(header_name=header_name, number_format=number_format, offset=offset)
            self.make_summary = self.Summary('', None)
            self.model_summary = self.Summary('', None)

        class Summary:
            def __init__(self, formula: str, formula_ranges: (tuple, ), formatting_rule: formatting.rule.Rule=None):
                """
                The formula to bring up a value in the summary
                :param formula: the formula to bring a value
                :param formula_ranges: a tuple with the up-most cells of vehicles to fall in the formula range relative
                  to the summary cell
                """
                self.formula = formula
                self.formula_ranges = formula_ranges
                self.formatting_rule = formatting_rule

            def mount(self, cur_col: int, cur_row: int, rows_amount: int, veh_mark_up_col: int=0):
                if '{0}' not in self.formula:
                    return None
                if '{x}' in self.formula:
                    formula = self.formula.replace('{x}', '{}:{}'.format(
                                                   '{}{}'.format(utils.get_column_letter(xl(veh_mark_up_col)),
                                                                 xl(cur_row + self.formula_ranges[0][0])),
                                                   '{}{}'.format(utils.get_column_letter(xl(veh_mark_up_col)),
                                                                 xl(cur_row + self.formula_ranges[0][0] +
                                                                    rows_amount - 1))))
                else:
                    formula = self.formula
                return formula.format(*tuple([
                    '{}:{}'.format(  # formula cells range
                        '{}{}'.format(utils.get_column_letter(xl(cur_col + formula_range[1])),
                                      xl(cur_row + formula_range[0])),
                        '{}{}'.format(utils.get_column_letter(xl(cur_col + formula_range[1])),
                                      xl(cur_row + formula_range[0] + rows_amount - 1)))
                    for formula_range in self.formula_ranges]))

    POSITION = {
        'title_header_row': 0,
        'time_header_row': 1,
        'info_header_row': 2,
        'first_sample_row': 3,
        'vehicle_col': 0,
        'prod_model_year_col': 1,
        'first_sample_col': 2,
        'summary_make_col': 0,
        'summary_time_header_row': 0
    }

    @abc.abstractmethod
    def generate_report(self):
        raise NotImplementedError()

    def write_headers_to_matrix(self, report_name: str):
        self.matrix[self.POSITION['title_header_row']][self.POSITION['vehicle_col']] = 'Data from {} to {}'.format(
            str(min(self.sample_dates)), str(max(self.sample_dates)))
        self.matrix[self.POSITION['time_header_row']][self.POSITION['vehicle_col']] = 'Vehicle'
        self.matrix[self.POSITION['time_header_row']][self.POSITION['prod_model_year_col']] = 'MY'
        self.matrix[self.POSITION['title_header_row']][self.POSITION['first_sample_col']] = report_name
        col = self.POSITION['first_sample_col']
        from datetime import datetime
        for sample_date in self.sample_dates:
            self.matrix[self.POSITION['time_header_row']][col] = \
                datetime.strptime(str(sample_date), '%Y%m%d').strftime('%B')
            for header in self.sample_headers:
                self.matrix[self.POSITION['info_header_row']][col] = header.header_name
                col += 1

    def fill_empty_vehicle_cells(self, version_row: int):
        """
        Fill the empty cells of the specified row so the analyst can see that those cells were looked for
        :param version_row:
        :return:
        """
        for sample_date_index in range(0, len(self.sample_dates)):
            if not self.matrix[version_row][self.POSITION['first_sample_col'] +
                                            len(self.sample_headers) * sample_date_index]:
                for header in self.sample_headers:
                    self.matrix[version_row][self.POSITION['first_sample_col']
                                             + len(self.sample_headers) * sample_date_index + header.offset] = '?'

    def write_make_header(self, make_name: str, amount_of_distinct_models_of_make: int,
                          amount_of_distinct_vehicles_of_make: int, make_header_row: int):
        self.matrix[make_header_row][self.POSITION['vehicle_col']] = make_name
        for sample_date_index in range(len(self.sample_dates)):
            first_column_of_sample = self.POSITION['first_sample_col'] + sample_date_index * len(self.sample_headers)
            for header in self.sample_headers:
                absolute_column_of_header = first_column_of_sample + header.offset
                self.matrix[make_header_row][absolute_column_of_header] = header.make_summary.mount(
                    cur_col=absolute_column_of_header,
                    cur_row=make_header_row,
                    rows_amount=amount_of_distinct_vehicles_of_make + amount_of_distinct_models_of_make,
                    veh_mark_up_col=self.vehicle_desc_mark_up_col)
                if header.make_summary.formatting_rule:
                    self.ws.conditional_formatting.add('{}{}'.format(
                        utils.get_column_letter(xl(absolute_column_of_header)),
                        xl(make_header_row)), header.make_summary.formatting_rule)

    def write_model_header(self, model_name: str, amount_of_distinct_vehicles_of_model: int, model_header_row: int):
        self.matrix[model_header_row][self.POSITION['vehicle_col']] = model_name
        self.matrix[model_header_row][self.vehicle_desc_mark_up_col] = 'm'
        for sample_date_index in range(len(self.sample_dates)):
            first_column_of_sample = self.POSITION['first_sample_col'] + sample_date_index * len(self.sample_headers)
            for header in self.sample_headers:
                absolute_column_of_header = first_column_of_sample + header.offset
                self.matrix[model_header_row][absolute_column_of_header] = header.model_summary.mount(
                    cur_col=absolute_column_of_header,
                    cur_row=model_header_row,
                    rows_amount=amount_of_distinct_vehicles_of_model)
                if header.model_summary.formatting_rule:
                    self.ws.conditional_formatting.add('{}{}'.format(
                        utils.get_column_letter(xl(absolute_column_of_header)),
                        xl(model_header_row)), header.model_summary.formatting_rule)

    def finish_worksheet(self, last_row_index: int):
        def style_range(cell_range, border=styles.Border(), fill=None, font=None, alignment=None):
            """
            Apply styles to a range of cells as if they were a single cell.

            :param cell_range: An excel range to style (e.g. A1:F20)
            :param border: An openpyxl Border
            :param fill: An openpyxl PatternFill or GradientFill
            :param font: An openpyxl Font object
            :param alignment: An openpyxl Alignment object
            """

            top = styles.Border(top=border.top)
            left = styles.Border(left=border.left)
            right = styles.Border(right=border.right)
            bottom = styles.Border(bottom=border.bottom)

            first_cell = self.ws[cell_range.split(":")[0]]
            if alignment:
                self.ws.merge_cells(cell_range)
                first_cell.alignment = alignment

            rows = self.ws[cell_range]
            if font:
                first_cell.font = font

            for cell in rows[0]:
                cell.border = cell.border + top
            for cell in rows[-1]:
                cell.border = cell.border + bottom

            for row in rows:
                l = row[0]
                r = row[-1]
                l.border = l.border + left
                r.border = r.border + right
                if fill:
                    for c in row:
                        c.fill = fill

        def add_header_xl_styles():
            """
            Merges, adds color, font and size to the report headers
            :return:
            """
            def title_header(style: styles.NamedStyle):
                """
                Add style to report title cells
                :param style:
                :return:
                """
                title_header_range = '{}:{}'.format(
                    '{}{}'.format(utils.get_column_letter(xl(self.POSITION['first_sample_col'])),
                                  xl(self.POSITION['title_header_row'])),
                    '{}{}'.format(utils.get_column_letter(xl(self.POSITION['first_sample_col'] +
                                                          len(self.sample_dates) * len(self.sample_headers))),
                                  xl(self.POSITION['title_header_row'])))
                self.ws.merge_cells(title_header_range)
                style_range(title_header_range, border=style.border, fill=style.fill,
                            font=style.font, alignment=style.alignment)

            def time_headers(style: styles.NamedStyle):
                """
                Add style to time headers
                :param style:
                :return:
                """
                vehicle_header_range = '{}:{}'.format(
                    '{}{}'.format(utils.get_column_letter(xl(self.POSITION['vehicle_col'])),
                                  xl(self.POSITION['time_header_row'])),
                    '{}{}'.format(utils.get_column_letter(xl(self.POSITION['vehicle_col'])),
                                  xl(self.POSITION['info_header_row'])))
                self.ws.merge_cells(vehicle_header_range)
                style_range(vehicle_header_range, border=style.border, fill=style.fill,
                            font=style.font, alignment=style.alignment)

                prod_model_year_header_range = '{}:{}'.format(
                    '{}{}'.format(utils.get_column_letter(xl(self.POSITION['prod_model_year_col'])),
                                  xl(self.POSITION['time_header_row'])),
                    '{}{}'.format(utils.get_column_letter(xl(self.POSITION['prod_model_year_col'])),
                                  xl(self.POSITION['info_header_row'])))
                self.ws.merge_cells(prod_model_year_header_range)
                style_range(prod_model_year_header_range, border=style.border, fill=style.fill,
                            font=style.font, alignment=style.alignment)

                max_sample_header_offset = max(self.sample_headers, key=lambda d: d.offset).offset
                for sample_date_index in range(len(self.sample_dates)):
                    cur_col_index = xl(self.POSITION['first_sample_col'] +
                                       sample_date_index * (max_sample_header_offset + 1))
                    time_header_range = '{}:{}'.format(
                        '{}{}'.format(utils.get_column_letter(cur_col_index),
                                      xl(self.POSITION['time_header_row'])),
                        '{}{}'.format(utils.get_column_letter(cur_col_index + max_sample_header_offset),
                                      xl(self.POSITION['time_header_row'])))
                    self.ws.merge_cells(time_header_range)
                    style_range(time_header_range, border=style.border, fill=style.fill,
                                font=style.font, alignment=style.alignment)

            def info_headers(style: styles.NamedStyle):
                """
                Add style to info headers
                :param style:
                :return:
                """
                info_headers_row = str(xl(self.POSITION['info_header_row']))
                for info_header_col in range(xl(self.POSITION['first_sample_col']), xl(self.vehicle_desc_mark_up_col)):
                    info_header_cell = '{}{}'.format(utils.get_column_letter(info_header_col), info_headers_row)
                    self.ws[info_header_cell].style = style

            title_header_style = styles.NamedStyle(name='title_header', font=styles.Font(sz=28),
                                                   alignment=styles.Alignment(horizontal='center', vertical='center'))
            time_header_style = styles.NamedStyle(name='time_header',
                                                  font=styles.Font(sz=12, b=True, color=styles.Color('F1F2F2')),
                                                  alignment=styles.Alignment(horizontal='center', vertical='center'),
                                                  border=styles.Border(bottom=styles.Side(color='F1F2F2', style='thin'),
                                                                       left=styles.Side(color='F1F2F2', style='thin')),
                                                  fill=styles.PatternFill(patternType='solid',
                                                                          fgColor=styles.Color('930004')))
            info_header_style = styles.NamedStyle(name='info_headers',
                                                  font=styles.Font(sz=12, b=True, color=styles.Color('F1F2F2')),
                                                  alignment=styles.Alignment(horizontal='center', vertical='center'),
                                                  border=styles.Border(bottom=styles.Side(color='F1F2F2', style='thin'),
                                                                       left=styles.Side(color='F1F2F2', style='thin')),
                                                  fill=styles.PatternFill(patternType='solid',
                                                                          fgColor=styles.Color('930004')))
            title_header(title_header_style)
            time_headers(time_header_style)
            info_headers(info_header_style)

        def add_content_xl_styles():
            """
            Add styles to content cells
            :return:
            """
            def apply_styles_to_rows():
                def swap_version_row_color():
                    if cur_version_style == light_version_row_style:
                        return dark_version_row_style
                    else:
                        return light_version_row_style

                cur_version_style = light_version_row_style
                veh_col_letter = utils.get_column_letter(xl(self.POSITION['vehicle_col']))
                prod_mdl_yr_col_letter = utils.get_column_letter(xl(self.POSITION['prod_model_year_col']))

                for row in range(self.POSITION['first_sample_row'], xl(last_row_index)):
                    if self.matrix[row][self.vehicle_desc_mark_up_col] == 'v':  # version row
                        if self.matrix[row][self.POSITION['vehicle_col']] != '':
                            cur_version_style = swap_version_row_color()
                        cur_style = cur_version_style
                    elif self.matrix[row][self.vehicle_desc_mark_up_col] == 'm':  # model row
                        cur_style = model_row_style
                    else:  # make row
                        cur_style = make_row_style

                    self.ws['{}{}'.format(veh_col_letter, xl(row))].style = cur_style
                    self.ws['{}{}'.format(prod_mdl_yr_col_letter, xl(row))].style = cur_style
                    sample_headers_amount = len(self.sample_headers)
                    for sample_date_index in range(len(self.sample_dates)):
                        for sample_header in self.sample_headers:
                            cell = '{}{}'.format(utils.get_column_letter(xl(
                                self.POSITION['first_sample_col']
                                + sample_headers_amount * sample_date_index
                                + sample_header.offset)), xl(row))
                            self.ws[cell].style = cur_style
                            self.ws[cell].number_format = sample_header.number_format

            make_row_style = styles.NamedStyle(name='make_row',
                                               font=styles.Font(sz=10, b=True, color=styles.Color('F1F2F2')),
                                               fill=styles.PatternFill(patternType='solid',
                                                                       fgColor=styles.Color('000000')))
            model_row_style = styles.NamedStyle(name='model_row',
                                                font=styles.Font(sz=10, b=True, color=styles.Color('000000')),
                                                fill=styles.PatternFill(patternType='solid',
                                                                        fgColor=styles.Color('939598')))
            light_version_row_style = styles.NamedStyle(name='light_version_row',
                                                        font=styles.Font(sz=10, b=True, color=styles.Color('000000')),
                                                        fill=styles.PatternFill(patternType='solid',
                                                                                fgColor=styles.Color('F1F2F2')))
            dark_version_row_style = styles.NamedStyle(name='dark_version_row',
                                                       font=styles.Font(sz=10, b=True, color=styles.Color('000000')),
                                                       fill=styles.PatternFill(patternType='solid',
                                                                               fgColor=styles.Color('DCDDDE')))
            apply_styles_to_rows()

        def add_overall_xl_styles():
            """
            Final alignments, border, split and views freezes
            :return:
            """
            self.ws.freeze_panes = self.ws['{}{}'.format(
                utils.get_column_letter(xl(self.POSITION['first_sample_col'])),
                str(xl(self.POSITION['first_sample_row'])))]
            self.ws.column_dimensions[utils.get_column_letter(xl(self.vehicle_desc_mark_up_col))].hidden = True

        add_header_xl_styles()
        add_content_xl_styles()
        add_overall_xl_styles()
