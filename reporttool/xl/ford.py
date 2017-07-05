from openpyxl import utils, formatting
from openpyxl.chart import (LineChart, Reference)
from openpyxl.comments import Comment

from . import base


class EquipmentAndIncentives(base.EvolutionReport):
    class Entity:
        class Incentive:
            def __init__(self, inc_data: tuple):
                self.code, self.jato_value, self.take_rate, self.manuf_contrib, self.interest_perc,\
                    self.deposit_perc, self.max_term, self.final_balance_perc, self.start_date, self.end_date,\
                    self.internal_comments, self.public_notes = inc_data

        class Price:
            def __init__(self, price_data: tuple):
                self.msrp, self.data_date = price_data

        class Volume:
            def __init__(self, vol_data: tuple):
                self.volume, self.data_date = vol_data

        class Equipment:
            def __init__(self, equip_data: tuple):
                self.equipment_value, self.data_date = equip_data

        def __init__(self, veh_data: tuple=None, grouped_data: list=None):
            if veh_data is None and grouped_data is None:
                self.make, self.model, self.version, self.prod_yr, self.model_yr = ['' for _ in range(5)]
                self.prices = []
                self._incentives = []
                self.volumes = []
                self.equipments = []
                return
            self.make, self.model, self.version, self.prod_yr, self.model_yr = veh_data
            self.prices = [self.Price(price_data=unique_price)
                           for unique_price in set([price[5:7] for price in grouped_data])]
            self._incentives = [self.Incentive(inc_data=unique_incentive)
                                for unique_incentive in set([incentive[7:19] for incentive in grouped_data])]
            self.volumes = [self.Volume(vol_data=unique_volume)
                            for unique_volume in set([volume[19:21] for volume in grouped_data])]
            self.equipments = [self.Equipment(equip_data=unique_equip)
                               for unique_equip in set([equipment[21:23] for equipment in grouped_data])]

        @property
        def pymy(self):
            return '{}/{}'.format(str(self.prod_yr % 100) if self.prod_yr else '?',
                                  str(self.model_yr % 100) if self.model_yr else '?')

        def get_oem_incentive(self, sample_date: int):
            def comment(incentive: EquipmentAndIncentives.Entity.Incentive):
                if incentive.code[:3] in ('SFS', 'SST', 'SSA'):
                    return incentive.internal_comments
                else:
                    return incentive.public_notes

            oem_incs = [oem_inc for oem_inc in self._incentives if oem_inc.code[:3]
                        in ('SFS', 'SST', 'SSA', 'DPR', 'DSO', 'OFL', 'OIN', 'ORA', 'OSV', 'OVT', 'OWY', 'OPA',
                            'OSO', 'OSA', 'PAS', 'PFR', 'POT', 'PRP', 'PSE', 'PSO', 'PSA', 'DOV', 'DLT', 'DTS')
                        and oem_inc.jato_value is not None
                        and oem_inc.start_date // 100 <= sample_date // 100 <= oem_inc.end_date // 100]
            if len(oem_incs) == 0:
                return '-', None
            elif len(oem_incs) == 1:
                return oem_incs[0].jato_value, comment(oem_incs[0])
            else:
                max_oem_inc = max(oem_incs, key=lambda oem_inc: oem_inc.jato_value)
                return max_oem_inc.jato_value, comment(max_oem_inc)

        def get_finance_incentive(self, sample_date: int):
            def comment(incentive: EquipmentAndIncentives.Entity.Incentive):
                return 'Taxa {0}%/{1}%/{2}x {3}Take Rate: {4}% x {5}'.format(
                    '?' if incentive.interest_perc is None else '{0:.2f}'.format(incentive.interest_perc),
                    '?' if incentive.deposit_perc is None else '{0:.2f}'.format(incentive.deposit_perc),
                    '?' if incentive.max_term is None else '{0:.2f}'.format(incentive.max_term),
                    '+ {0:.2f}% (Balão) '.format(incentive.final_balance_perc) if incentive.final_balance_perc > 0
                    else ' ',
                    '{0:.2f}'.format(incentive.take_rate),
                    '{0:.2f}'.format(incentive.jato_value))

            finances = [finance for finance in self._incentives if finance.code.startswith('F')
                        and finance.jato_value is not None and finance.take_rate is not None
                        and finance.start_date // 100 <= sample_date // 100 <= finance.end_date // 100]
            if len(finances) == 0:
                return '-', None
            if len(finances) == 1:
                    return finances[0].jato_value / finances[0].take_rate, comment(finances[0])
            else:
                finances.sort(key=lambda f: f.jato_value)
                if len(finances) > 2 and finances[1].jato_value - finances[2].jato_value >= 500:
                    return finances[0].jato_value / finances[0].take_rate, ' | '.join(
                        (comment(finances[0]), comment(finances[1]), comment(finances[2])))
                else:
                    return finances[0].jato_value / finances[0].take_rate, ' | '.join(
                        (comment(finances[0]), comment(finances[1])))

    def __init__(self, data: tuple):
        def group_data():
            from itertools import groupby
            from operator import itemgetter
            return [self.Entity(veh_data=veh_data[:], grouped_data=list(grouped_data)) for veh_data, grouped_data
                    in groupby(data, key=itemgetter(0, 1, 2, 3, 4))]

        def define_sample_dates(veh_sample_dates: set, vol_sample_dates: set):
            max_vol_sample_date = max(vol_sample_dates)
            sample_dates = [veh_sample_date for veh_sample_date in veh_sample_dates
                            if veh_sample_date > max_vol_sample_date]
            sample_dates.extend(vol_sample_dates)
            return sorted(sample_dates)

        def create_header(name: str, number_format: str, offset,
                          make_summary_formula: str, make_summary_formula_ranges: (tuple, ),
                          model_summary_formula: str, model_summary_formula_ranges: (tuple,),
                          make_summary_format_rule: formatting.Rule=None,
                          model_summary_format_rule: formatting.Rule=None):
            header = self.Header(header_name=name, number_format=number_format, offset=offset)
            header.make_summary = header.Summary(make_summary_formula,
                                                 make_summary_formula_ranges, make_summary_format_rule)
            header.model_summary = header.Summary(model_summary_formula,
                                                  model_summary_formula_ranges, model_summary_format_rule)
            return header

        def create_net_price_icon_set():
            return formatting.rule.IconSet('3Arrows',
                                           cfvo=[formatting.rule.FormatObject(type='num', val=0),
                                                 formatting.rule.FormatObject(type='num', val=0, gte=True),
                                                 formatting.rule.FormatObject(type='num', val=0, gte=False)],
                                           showValue=True)

        super().__init__()
        self.sample_dates = define_sample_dates(
            veh_sample_dates=set((row[6] // 100) * 100 + 1 for row in data),
            vol_sample_dates=set(row[-3] for row in data if row[-3] is not None))
        self.entities = group_data()
        data = None
        if not self.entities:
            raise IndexError('Cant generate report without data')
        self.sample_headers.append(
            create_header(name='MSRP', number_format='#,###', offset=0,
                          make_summary_formula='=IF(ISERROR(AVERAGEIF({x},"m",{0})),"-",AVERAGEIF({x},"m",{0}))',
                          make_summary_formula_ranges=[(1, 0), ],
                          model_summary_formula='=IF(ISERROR(AVERAGE({0})),"-",AVERAGE({0}))',
                          model_summary_formula_ranges=[(1, 0)]))
        self.sample_headers.append(
            create_header(name='Equip. Value', number_format='#,###', offset=1,
                          make_summary_formula='=IF(ISERROR(SUMIF({x},"m",{0})),"-",SUMIF({x},"m",{0}))',
                          make_summary_formula_ranges=[(1, 0), ],
                          model_summary_formula='=IF(ISERROR(SUM({0})),"-",SUM({0}))',
                          model_summary_formula_ranges=[(1, 0)]))
        self.sample_headers.append(
            create_header(name='Manuf. Contrib.', number_format='#,###', offset=2,
                          make_summary_formula='=IF(ISERROR(AVERAGEIF({x},"m",{0})),"-",AVERAGEIF({x},"m",{0}))',
                          make_summary_formula_ranges=[(1, 0), ],
                          model_summary_formula='=IF(ISERROR(AVERAGE({0})),"-",AVERAGE({0}))',
                          model_summary_formula_ranges=[(1, 0)]))
        self.sample_headers.append(
            create_header(name='Finance Incentive', number_format='#,###', offset=3,
                          make_summary_formula='=IF(ISERROR(AVERAGEIF({x},"m",{0})),"-",AVERAGEIF({x},"m",{0}))',
                          make_summary_formula_ranges=[(1, 0), ],
                          model_summary_formula='=IF(ISERROR(AVERAGE({0})),"-",AVERAGE({0}))',
                          model_summary_formula_ranges=[(1, 0)]))
        self.sample_headers.append(
            create_header(name='Net Price', number_format='#.00%', offset=4,
                          make_summary_formula='=IF(ISERROR(AVERAGEIF({x},"m",{0})),"-",AVERAGEIF({x},"m",{0}))',
                          make_summary_formula_ranges=[(1, 0)],
                          make_summary_format_rule=formatting.Rule(
                              type='iconSet',
                              iconSet=create_net_price_icon_set()),
                          model_summary_formula='=IF(ISERROR(AVERAGE({0})),"-",AVERAGE({0}))',
                          model_summary_formula_ranges=[(1, 0)],
                          model_summary_format_rule=formatting.Rule(
                              type='iconSet',
                              iconSet=create_net_price_icon_set())))
        self.sample_headers.append(
            create_header(name='Volume', number_format='#,###', offset=5,
                          make_summary_formula='=IF(ISERROR(SUMIF({x},"m",{0})),"-",SUMIF({x},"m",{0}))',
                          make_summary_formula_ranges=[(1, 0), ],
                          model_summary_formula='=IF(ISERROR(SUM({0})),"-",SUM({0}))',
                          model_summary_formula_ranges=[(1, 0)]))
        self.vehicle_desc_mark_up_col = \
            self.POSITION['first_sample_col'] + len(self.sample_dates) * len(self.sample_headers)

    def __str__(self):
        from datetime import datetime
        return '{}.{}_{}'.format(super(EquipmentAndIncentives, self).__str__(),
                                 'Equipment_and_Incentives_Report',
                                 datetime.today().strftime('%Y%m%d'))

    def generate_report(self):
        def size_matrix_rows():
            makes_amount = len(set(veh.make for veh in self.entities))
            models_amount = len(set(veh.model for veh in self.entities))
            return self.POSITION['first_sample_row'] + makes_amount + models_amount + len(self.entities)

        def on_make_change(new_make_name: str):
            if temp_ent.make:
                self.fill_empty_vehicle_cells(version_row=row_index)
            temp_ent.make = new_make_name
            temp_ent.model = None
            vehicles_of_make = [veh for veh in self.entities if veh.make == new_make_name]
            models_of_make = set(veh.model for veh in vehicles_of_make)
            self.write_make_header(make_name=new_make_name,
                                   amount_of_rows_for_make=len(vehicles_of_make) + len(models_of_make),
                                   make_header_row=row_index + 1)

        def on_model_change(new_model_name: str):
            if temp_ent.model:
                self.fill_empty_vehicle_cells(version_row=row_index)
            temp_ent.model = new_model_name
            temp_ent.version = None
            amount_of_rows_for_model = len(
                set([(veh.version, veh.prod_yr, veh.model_yr)
                     for veh in self.entities if veh.make == entity.make and veh.model == entity.model]))
            self.write_model_header(model_name=new_model_name,
                                    amount_of_rows_for_model=amount_of_rows_for_model,
                                    model_header_row=row_index + 1)

        def on_version_change(new_version_name: str):
            if temp_ent.version:
                self.fill_empty_vehicle_cells(version_row=row_index)
            temp_ent.version = new_version_name
            temp_ent.prod_yr = None
            temp_ent.model_yr = None
            self.matrix[row_index + 1][self.vehicle_desc_mark_up_col] = 'v'
            self.matrix[row_index + 1][self.POSITION['vehicle_col']] = new_version_name

        def on_model_year_change(new_pymy: str):
            if temp_ent.pymy != '?/?':
                self.fill_empty_vehicle_cells(version_row=row_index)
            temp_ent.prod_yr = entity.prod_yr
            temp_ent.model_yr = entity.model_yr
            self.matrix[row_index + 1][self.vehicle_desc_mark_up_col] = 'v'
            self.matrix[row_index + 1][self.POSITION['prod_model_year_col']] = new_pymy

        def write_vehicle_data(vehicle_ent: self.Entity, version_row: int):
            for idx, sample_date in enumerate(self.sample_dates):
                column_index = self.POSITION['first_sample_col'] + idx * len(self.sample_headers)
                self.matrix[version_row][column_index] = next(
                    (p.msrp for p in vehicle_ent.prices if p.data_date == max(
                        vehicle_ent.prices, key=lambda p: (p.data_date // 100) * 100 + 1 <= sample_date).data_date),
                    '?')
                self.matrix[version_row][column_index + 1] = next(
                    (e.equipment_value for e in vehicle_ent.equipments if e.data_date == sample_date), '?')

                oem_incentive = vehicle_ent.get_oem_incentive(sample_date=sample_date)
                self.matrix[version_row][column_index + 2] = oem_incentive[0]
                if oem_incentive[1]:
                    self.ws['{}{}'.format(utils.get_column_letter(base.xl(column_index + 2)),
                                          str(base.xl(version_row)))].comment = Comment(text=oem_incentive[1],
                                                                                        author='ReportTool')

                finance_incentive = vehicle_ent.get_finance_incentive(sample_date=sample_date)
                self.matrix[version_row][column_index + 3] = finance_incentive[0]
                if finance_incentive[1]:
                    self.ws['{}{}'.format(utils.get_column_letter(base.xl(column_index + 3)),
                                          str(base.xl(version_row)))].comment = Comment(text=finance_incentive[1],
                                                                                        author='ReportTool')
                if idx > 0:
                    self.matrix[version_row][column_index + 4] = \
                        '=IF(AND(ISNUMBER({1}{0}),ISNUMBER({4}{0})),' \
                        '({1}{0}-SUM({2}{0}:{3}{0}))/({4}{0}-SUM({5}{0}:{6}{0}))-1,"-")'.format(
                            str(base.xl(version_row)),
                            utils.get_column_letter(base.xl(column_index)),
                            utils.get_column_letter(base.xl(column_index + 1)),
                            utils.get_column_letter(base.xl(column_index + 3)),
                            utils.get_column_letter(base.xl(column_index - 6)),
                            utils.get_column_letter(base.xl(column_index - 5)),
                            utils.get_column_letter(base.xl(column_index - 3)))
                else:
                    self.matrix[version_row][column_index + 4] = '-'

                self.ws.conditional_formatting.add('{}{}'.format(
                    utils.get_column_letter(base.xl(column_index + 4)),
                    base.xl(version_row)), self.sample_headers[4].model_summary.formatting_rule)
                self.matrix[version_row][column_index + 5] = next(
                    (v.volume for v in vehicle_ent.volumes if v.data_date == sample_date), '?')

        self.ws.title = 'Comparative'
        temp_ent = self.Entity()
        row_index = self.POSITION['first_sample_row'] - 1
        self.matrix = [['' for _ in range(self.vehicle_desc_mark_up_col + 1)] for _ in range(size_matrix_rows())]
        self.write_headers_to_matrix('Ford - Equipments & Incentives')
        for entity in self.entities:
            if temp_ent.make != entity.make:
                on_make_change(new_make_name=entity.make)
                row_index += 1
            if temp_ent.model != entity.model:
                on_model_change(new_model_name=entity.model)
                row_index += 1
            if temp_ent.version != entity.version:
                on_version_change(new_version_name=entity.version)
            if temp_ent.pymy != entity.pymy:
                on_model_year_change(new_pymy=entity.pymy)
                row_index += 1
            write_vehicle_data(vehicle_ent=entity, version_row=row_index)

        self.fill_empty_vehicle_cells(version_row=row_index)
        self.write_matrix_to_xl()
        self.finish_worksheet(last_row_index=row_index)
        self.create_summary_chart(last_row=row_index)
        self.write_to_disc()

    def create_summary_chart(self, last_row: int):
        def create_data_table():
            from datetime import datetime

            def extract_summary_data(summary_row: int, header_name: str):
                try:
                    net_price_header = [header for header in self.sample_headers
                                        if header.header_name == header_name][0]
                except IndexError:
                    raise IndexError('Could not find {} header'.format(header_name))

                max_header_offset = max(self.sample_headers, key=lambda h: h.offset).offset

                summary_list = [self.matrix[summary_row][self.POSITION['vehicle_col']], ]
                for date_index, _ in enumerate(self.sample_dates):
                    summary_list.append('=Comparative!{}{}'.format(
                        utils.get_column_letter(
                            base.xl(self.POSITION['first_sample_col']
                                    + net_price_header.offset
                                    + date_index * (max_header_offset + 1))),
                        base.xl(summary_row)))

                return summary_list

            mark_up_col_letter = utils.get_column_letter(base.xl(self.vehicle_desc_mark_up_col))
            make_summary_list = [[]]

            make_summary_list[0].append('Date')
            for date in self.sample_dates:
                make_summary_list[0].append(datetime.strptime(str(date), '%Y%m%d').strftime('%Y-%m'))

            for row in range(self.POSITION['first_sample_row'], last_row + 1):
                if self.ws['{}{}'.format(mark_up_col_letter, base.xl(row))].value not in ('m', 'v'):
                    make_summary_list.append(extract_summary_data(summary_row=row, header_name='Net Price'))

            # Transpose
            make_summary_list = list(map(list, zip(*make_summary_list)))

            for make_summary in make_summary_list:
                chart_sheet.append(make_summary)

        chart_sheet = self.wb.create_sheet(title='Chart')
        create_data_table()

        chart = LineChart()
        chart.title = 'Make Summary'
        chart.style = 11
        chart.y_axis.title = 'Net Price'
        chart.y_axis.number_format = '#.00%'
        chart.x_axis.title = 'Date'
        chart.x_axis.number_format = 'yyyy-mm'

        data = Reference(chart_sheet, min_col=2, min_row=1,
                         max_col=len(set(veh.make for veh in self.entities))+1, max_row=len(self.sample_dates)+1)
        chart.add_data(data, titles_from_data=True)
        dates = Reference(chart_sheet, min_col=1, min_row=2, max_row=len(self.sample_dates)+1)
        chart.set_categories(dates)

        chart_sheet.add_chart(chart, 'A{}'.format(base.xl(2 + len(self.sample_dates))))
