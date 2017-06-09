from openpyxl import utils
from openpyxl import formatting
from openpyxl.comments import Comment
from . import _base


class EquipmentAndIncentives(_base.EvolutionReport):
    class Entity:
        class Incentive:
            def __init__(self, inc_data: tuple):
                self.code, self.jato_value, self.take_rate, self.manuf_contrib, self.interest_perc, \
                self.deposit_perc, self.max_term, self.internal_comments, self.public_notes = inc_data

        def __init__(self, veh_data: tuple, incentives: list):
            self.make, self.model, self.version, self.prod_yr, self.model_yr, self.msrp, self.sample_date,\
                self.volume, self.value = veh_data
            self._incentives = [self.Incentive(inc_data=incentive[10:]) for incentive in incentives]

        @property
        def pymy(self):
            return '{}/{}'.format(str(self.prod_yr % 100) if self.prod_yr else '?',
                                  str(self.model_yr % 100) if self.model_yr else '?')

        @property
        def oem_incentive(self):
            oem_incs = [oem_inc for oem_inc in self._incentives if oem_inc.code[:3]
                        in ('DAT', 'DLT', 'DOV', 'DPR', 'DSO', 'DTS', 'SFS', 'SSA', 'SST')]
            if len(oem_incs) == 0:
                return '-', None
            elif len(oem_incs) == 1:
                return oem_incs[0].jato_value, oem_incs[0].public_notes
            else:
                max_oem_inc = max(oem_incs, key=lambda oem_inc: oem_inc.jato_value)
                return max_oem_inc.jato_value, max_oem_inc.public_notes

        @property
        def finance_incentive(self):
            finances = [finance for finance in self._incentives if finance.code.startswith('F')]
            if len(finances) == 0:
                return '-', None
            if len(finances) == 1:
                return finances[0].jato_value / finances[0].take_rate, None
            else:
                finances.sort(key=lambda f: f.jato_value / f.take_rate)
                return (finances[0].jato_value / finances[0].take_rate,
                        ('Taxa {0:.2f}%/ {1}%/ {2}x Take Rate sum: {3}'.format(finances[1].interest_perc,
                                                                               finances[1].deposit_perc,
                                                                               finances[1].max_term,
                                                                               sum(f.take_rate for f in finances))))

    def __init__(self, data: tuple):
        def group_data():
            from itertools import groupby
            from operator import itemgetter
            return [self.Entity(veh_data=veh_data[1:], incentives=incentives) for veh_data, incentives
                    in groupby(data, key=itemgetter(0, 1, 2, 3, 4, 5, 6, 7, 8, 9))]

        def create_header(name: str, number_format: str, offset,
                          make_summary_formula: str, make_summary_formula_ranges: (tuple, ),
                          model_summary_formula: str, model_summary_formula_ranges: (tuple,),
                          make_summary_format_rule: formatting.Rule=None,
                          model_summary_format_rule: formatting.Rule=None):
            header = self.Header(header_name=name, number_format=number_format, offset=offset)
            header.make_summary = header.Summary(make_summary_formula,
                                                 make_summary_formula_ranges, make_summary_format_rule)
            header.model_summary = header.Summary(model_summary_formula,
                                                  model_summary_formula_ranges, model_summary_format_rule)
            return header

        def create_net_price_icon_set():
            return formatting.rule.IconSet('3Arrows',
                                           cfvo=[formatting.rule.FormatObject(type='num', val=-1),
                                                 formatting.rule.FormatObject(type='num', val=0),
                                                 formatting.rule.FormatObject(type='num', val=1)],
                                           showValue=True)

        super().__init__()
        self.entities = group_data()
        data = None
        if not self.entities:
            raise IndexError('Cant generate report without data')
        self.sample_dates = sorted(set(entity.sample_date for entity in self.entities))
        self.sample_headers.append(create_header(name='MSRP', number_format='#,###', offset=0,
                                                 make_summary_formula='=IF(ISERROR(AVERAGEIF({x},"m",{0})),"-",'
                                                                      'AVERAGEIF({x},"m",{0}))',
                                                 make_summary_formula_ranges=[(1, 0), ],
                                                 model_summary_formula='=IF(ISERROR(AVERAGE({0})),"-",AVERAGE({0}))',
                                                 model_summary_formula_ranges=[(1, 0)]))
        self.sample_headers.append(create_header(name='Equip. Value', number_format='#,###', offset=1,
                                                 make_summary_formula='=IF(ISERROR(SUMIF({x},"m",{0})),"-",'
                                                                      'SUMIF({x},"m",{0}))',
                                                 make_summary_formula_ranges=[(1, 0), ],
                                                 model_summary_formula='=IF(ISERROR(SUM({0})),"-",SUM({0}))',
                                                 model_summary_formula_ranges=[(1, 0)]))
        self.sample_headers.append(create_header(name='Manuf. Contrib.', number_format='#,###', offset=2,
                                                 make_summary_formula='=IF(ISERROR(AVERAGEIF({x},"m",{0})),"-",'
                                                                      'AVERAGEIF({x},"m",{0}))',
                                                 make_summary_formula_ranges=[(1, 0), ],
                                                 model_summary_formula='=IF(ISERROR(AVERAGE({0})),"-",AVERAGE({0}))',
                                                 model_summary_formula_ranges=[(1, 0)]))
        self.sample_headers.append(create_header(name='Finance Incentive', number_format='#,###', offset=3,
                                                 make_summary_formula='=IF(ISERROR(AVERAGEIF({x},"m",{0})),"-",'
                                                                      'AVERAGEIF({x},"m",{0}))',
                                                 make_summary_formula_ranges=[(1, 0), ],
                                                 model_summary_formula='=IF(ISERROR(AVERAGE({0})),"-",AVERAGE({0}))',
                                                 model_summary_formula_ranges=[(1, 0)]))
        self.sample_headers.append(create_header(name='Net Price', number_format='#,###', offset=4,
                                                 make_summary_formula='=IF(ISERROR(AVERAGEIF({x},"m",{0})-{1}),'
                                                                      'IF(ISERROR(AVERAGEIF({x},"m",{0})),"-",'
                                                                      'AVERAGEIF({x},"m",{0})),'
                                                                      'AVERAGEIF({x},"m",{0})-{1})',
                                                 make_summary_formula_ranges=[(1, 0), (1, -6)],
                                                 model_summary_formula='=IF(ISERROR(AVERAGE({0})),"-",AVERAGE({0}))',
                                                 model_summary_formula_ranges=[(1, 0)],
                                                 make_summary_format_rule=formatting.Rule(type='iconSet',
                                                                                          iconSet=
                                                                                          create_net_price_icon_set())))
        self.sample_headers.append(create_header(name='Volume', number_format='#,###', offset=5,
                                                 make_summary_formula='=IF(ISERROR(SUMIF({x},"m",{0})),"-",'
                                                                      'SUMIF({x},"m",{0}))',
                                                 make_summary_formula_ranges=[(1, 0), ],
                                                 model_summary_formula='=IF(ISERROR(SUM({0})),"-",SUM({0}))',
                                                 model_summary_formula_ranges=[(1, 0)]))
        self.vehicle_desc_mark_up_col = self.POSITION['first_sample_col'] \
                + len(self.sample_dates) * len(self.sample_headers)

    def __str__(self):
        from datetime import datetime
        return '{}.{}_{}'.format(super(EquipmentAndIncentives, self).__str__(),
                                 'Equipment_and_Incentives_Report',
                                 datetime.today().strftime('%Y%m%d'))

    def generate_report(self):
        def size_matrix_rows():
            makes_amount = len(set(entity.make for entity in self.entities))
            models_amount = len(set(entity.model for entity in self.entities))
            return self.POSITION['first_sample_row'] + makes_amount + models_amount + len(self.entities)

        def on_make_change(new_make_name: str):
            if temp_ent.make:
                self.fill_empty_vehicle_cells(version_row=row_index)
            temp_ent.make = new_make_name
            temp_ent.model = None
            vehicles_of_make = set([(veh.model, veh.version, veh.prod_yr, veh.model_yr) for veh in self.entities
                                    if veh.make == new_make_name])
            amount_of_models = len(set([veh[0] for veh in vehicles_of_make]))
            veh_versions_of_make = set([(veh[1], veh[2], veh[3]) for veh in vehicles_of_make])
            amount_of_veh_of_make = len([veh[0] for veh in veh_versions_of_make])
            self.write_make_header(make_name=new_make_name,
                                   amount_of_distinct_models_of_make=amount_of_models,
                                   amount_of_distinct_vehicles_of_make=amount_of_veh_of_make,
                                   make_header_row=row_index + 1)

        def on_model_change(new_model_name: str):
            if temp_ent.model:
                self.fill_empty_vehicle_cells(version_row=row_index)
            temp_ent.model = new_model_name
            temp_ent.version = None
            amount_of_rows_for_model = len(set([(veh.version, veh.prod_yr, veh.model_yr) for veh in self.entities
                                            if veh.make == entity.make and veh.model == entity.model]))
            self.write_model_header(model_name=new_model_name,
                                    amount_of_distinct_vehicles_of_model=amount_of_rows_for_model,
                                    model_header_row=row_index + 1)

        def on_version_change(new_version_name: str):
            if temp_ent.version:
                self.fill_empty_vehicle_cells(version_row=row_index)
            temp_ent.version = new_version_name
            temp_ent.prod_yr = None
            temp_ent.model_yr = None
            self.matrix[row_index + 1][self.vehicle_desc_mark_up_col] = 'v'
            self.matrix[row_index + 1][self.POSITION['vehicle_col']] = new_version_name

        def on_model_year_change(new_pymy: str):
            if temp_ent.pymy != '?/?':
                self.fill_empty_vehicle_cells(version_row=row_index)
            temp_ent.prod_yr = entity.prod_yr
            temp_ent.model_yr = entity.model_yr
            self.matrix[row_index + 1][self.vehicle_desc_mark_up_col] = 'v'
            self.matrix[row_index + 1][self.POSITION['prod_model_year_col']] = new_pymy

        def write_vehicle_data(vehicle_ent: self.Entity, version_row: int):
            sample_date_index = self.sample_dates.index(vehicle_ent.sample_date)
            column_index = self.POSITION['first_sample_col'] + sample_date_index * len(self.sample_headers)
            self.matrix[version_row][column_index] = vehicle_ent.msrp
            self.matrix[version_row][column_index + 1] = vehicle_ent.value if vehicle_ent.value else 0
            oem_incentive = vehicle_ent.oem_incentive
            self.matrix[version_row][column_index + 2] = oem_incentive[0]
            if oem_incentive[1]:
                self.ws['{}{}'.format(utils.get_column_letter(_base.xl(column_index + 2)),
                                      str(_base.xl(version_row)))].comment = Comment(text=oem_incentive[1],
                                                                                     author='ReportTool')
            finance_incentive = vehicle_ent.finance_incentive
            self.matrix[version_row][column_index + 3] = finance_incentive[0]
            if finance_incentive[1]:
                self.ws['{}{}'.format(utils.get_column_letter(_base.xl(column_index + 3)),
                                      str(_base.xl(version_row)))].comment = Comment(text=finance_incentive[1],
                                                                                     author='ReportTool')
            self.matrix[version_row][column_index + 4] = '={}-SUM({}:{})'.format(
                '{}{}'.format(utils.get_column_letter(_base.xl(column_index)), str(_base.xl(version_row))),
                '{}{}'.format(utils.get_column_letter(_base.xl(column_index + 1)), str(_base.xl(version_row))),
                '{}{}'.format(utils.get_column_letter(_base.xl(column_index + 3)), str(_base.xl(version_row))))
            self.matrix[version_row][column_index + 5] = vehicle_ent.volume

        temp_ent = self.Entity(veh_data=tuple([None for _ in range(0, 9)]), incentives=[])
        row_index = self.POSITION['first_sample_row'] - 1
        self.matrix = [['' for _ in range(self.vehicle_desc_mark_up_col + 1)] for _ in range(size_matrix_rows())]
        self.write_headers_to_matrix('Ford - Equipments & Incentives')
        for entity in self.entities:
            if temp_ent.make != entity.make:
                on_make_change(new_make_name=entity.make)
                row_index += 1
            if temp_ent.model != entity.model:
                on_model_change(new_model_name=entity.model)
                row_index += 1
            if temp_ent.version != entity.version:
                on_version_change(new_version_name=entity.version)
            if temp_ent.pymy != entity.pymy:
                on_model_year_change(new_pymy=entity.pymy)
                row_index += 1
            write_vehicle_data(vehicle_ent=entity, version_row=row_index)

        self.fill_empty_vehicle_cells(version_row=row_index)
        self.write_matrix_to_xl()
        self.finish_worksheet(last_row_index=row_index)
        self.write_to_disc()
